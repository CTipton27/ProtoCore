import shutil
from pathlib import Path
import sys
import subprocess
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

run_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
log_name = input("Please enter a name for the log directory: ")

# Grabs all directories needed to run scripts and log files.
root = Path(__file__).resolve().parent.parent
script_dir = root / "scripts"
tb_dir = root / "testbenches"
src_dir = root / "src"
log_dir = root / "logs"

current_log_dir = log_dir / log_name
current_log_dir.mkdir(exist_ok=True)

log_file = current_log_dir / f"{log_name}_{run_time}.log"
excel_file =  current_log_dir / f"{log_name}_{run_time}.xlsx"

#Dict allows for multiple sources/tbs
testbenches = {
    "alu": {
        "tb": tb_dir / "ALU_TB.v",
        "src_files": [src_dir / "ALU.v"],
        "xelab_opts": ["work.ALU_TB", "-s", "alu_tb_snapshot"],
        "xsim_opts": ["alu_tb_snapshot"],
        "log_file": "ALUlog.txt",
        "excel_log_file": "EALUlog.txt",
        "excel_format" : ["time", "a", "b", "opcode", "carry", "zero", "out", "fail"]
    },
    "reg_file": {
        "tb": tb_dir / "REG_TB.v",
        "src_files": [src_dir / "reg_file.v"],
        "xelab_opts": ["work.REG_TB", "-s", "reg_tb_snapshot"],
        "xsim_opts": ["reg_tb_snapshot"],
        "log_file": "REGlog.txt",
        "excel_log_file": "EREGlog.txt",
        "excel_format" : ["time", "ra", "rb", "wa", "wd", "we", "read_a", "read_b", "fail"]
    },
    "datapath": {
        "tb": tb_dir / "DATAPATH_TB.v",
        "src_files": [src_dir / "datapath.v", src_dir / "ALU.v", src_dir / "reg_file.v"],
        "xelab_opts": ["work.DATAPATH_TB", "-s", "datapath_tb_snapshot"],
        "xsim_opts": ["datapath_tb_snapshot"],
        "log_file": "DATAPATHlog.txt",
        "excel_log_file": "EDATAPATHlog.txt",
        "excel_format" : ["time", "ra_addr", "rb_addr", "write_addr", "top_data", "write_en", "read_a", "read_b", "alu_opcode", "alu_zero", "alu_carry", "alu_imm_flag", "status"]
    }
}

settings = Path("C:Xilinx/Vivado/2024.2/settings64.bat")

choice = input("Which testbench do you want to run? (alu, reg_file, datapath)? ").strip().lower()

if choice not in testbenches:
    print("Invalid choice")
    sys.exit(1)

#Builds the command - Starts by finding all sources and testbenches needed for xvlog
tb = testbenches[choice]["tb"]
sources = testbenches[choice]["src_files"]
all_files = [tb] + sources
cmd_files = " ".join(str(f) for f in all_files)

#builds xvlog, xelab, and xsim commands
xvlog_cmd = f"xvlog {cmd_files}"
xelab_cmd = "xelab " + " ".join(testbenches[choice]["xelab_opts"])
xsim_cmd = "xsim " + " ".join(testbenches[choice]["xsim_opts"]) + " -runall" #Prevents clock from stalling xsim


#joins full command for pass into subprocess.
full_cmd = f"{settings} && {xvlog_cmd} && {xelab_cmd} && {xsim_cmd}"

# Run and log output
with open(log_file, 'a') as f:
    result = subprocess.run(full_cmd, shell=True, stdout=f, stderr=subprocess.STDOUT)

if result.returncode != 0:
    print(f"Simulation failed with return code {result.returncode}")
    sys.exit(result.returncode)
else:
    print("Simulation completed successfully.")

# Extracts the plain log file from the scripts folder and moves it into logs
vivado_default_log = script_dir / testbenches[choice]["log_file"]
if vivado_default_log.exists():
    # Destination path for log
    dest_log_path = current_log_dir / testbenches[choice]["log_file"]
    try:
        # Moves file to log folder
        shutil.move(str(vivado_default_log), str(dest_log_path))
        print(f"Moved log file to {dest_log_path}")
    except Exception as e:
        print(f"Failed to move log file: {e}")
else:
    print(f"Expected log file {vivado_default_log} not found.")

# Extracts the plain Excel file from the scripts folder and moves it into logs
vivado_excel_log = script_dir / testbenches[choice]["excel_log_file"]
if vivado_excel_log.exists():
    dest_excel_path = current_log_dir / testbenches[choice]["excel_log_file"]
    try:
        # Moves file to log folder
        shutil.move(str(vivado_excel_log), str(dest_excel_path))
        print(f"Moved log file to {dest_excel_path}")
    except Exception as e:
        print(f"Failed to move excel log file: {e}")
else:
    print(f"Expected log file {vivado_default_log} not found.")

#removes unnecessary compilation files.
for item in script_dir.iterdir():
    if item.is_file():
        if item.suffix not in [".py", ".py~", ".asm"]:
            try:
                item.unlink()
            except Exception as e:
                print(f"Warning: Failed to delete file {item}: {e}")
    elif item.is_dir():
        if item.name not in [".idea", ".venv"]:
            try:
                shutil.rmtree(item)
            except Exception as e:
                print(f"Warning: Failed to delete directory {item}: {e}")
print("Cleanup complete.")

# Begins the formatting of the Excel plaintext into a .xlsx file
workbook = Workbook()
sheet = workbook.active
sheet.append(testbenches[choice]["excel_format"])

def try_convert(value):
    # Try int first
    try:
        return int(value)
    except ValueError:
        pass
    # If 'xx' or other strings that aren't numbers, keep as-is
    return value

# Write data rows
with open(current_log_dir / testbenches[choice]["excel_log_file"], 'r') as f:
    for line in f:
        line = line.strip().split(' | ')
        line = [try_convert(val) for val in line]
        if not line or len(line) != len(testbenches[choice]["excel_format"]):
            raise ValueError(
                f"Invalid excel line received: expected {len(testbenches[choice]['excel_format'])}, received {len(line)}. Line: {line}"
            )
        sheet.append(line)

last_col_index = len(testbenches[choice]["excel_format"])
last_col_letter = get_column_letter(last_col_index)
red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C", fill_type="solid")

start_row = 2 #row 1 is headers
end_row = sheet.max_row

range_str = f"A{start_row}:{last_col_letter}{end_row}"
fail_formula = f"${last_col_letter}2=1" #$ fixes column, but the row is variable for range.
success_formula = f"${last_col_letter}2=0"
neutral_formula = f"${last_col_letter}2=2"

fail_rule = FormulaRule(formula=[fail_formula], fill=red_fill)
success_rule = FormulaRule(formula=[success_formula], fill=green_fill)
neutral_rule = FormulaRule(formula=[neutral_formula], fill=yellow_fill)

sheet.conditional_formatting.add(range_str, fail_rule)
sheet.conditional_formatting.add(range_str, success_rule)
sheet.conditional_formatting.add(range_str, neutral_rule)

workbook.save(excel_file)